from cgitb import handler
from openpyxl import *
from random import *
from tkinter import *
from tkinter import filedialog
import tkinter.messagebox as msgbox

# length_word = 0
# groupe = []
# wb = load_workbook("D:\\user\\단어\\프랑스어 단어장.xlsx")
# ws = wb.active
# length_word =len(list(ws.rows))

# for word_row in ws.iter_rows(min_row=2, min_col=2, max_col=4, values_only=True):
#     groupe.append(word_row)
# print(groupe) #groupe이라는 리스트 자료의 0번째 인덱싱+0번째 값인 첫번째 튜플의 0번째 인덱싱
# shuffle(groupe)
# print(groupe)
        
groupe = []
length_word = 0
root = Tk()
root.title("test word")
root.geometry()


list_file = Listbox(root, selectmode="extended", height=0, width=30)
list_file.grid(row=0, column=0, columnspan=2)
list_file.insert(END, "테스트할 xlsx파일을 추가하세요")

mean_search = Entry(root, width=30)
mean_search.grid(row=3, column=0, columnspan=2)
mean_search.insert(END, "뜻을 입력하세요")    

word_search = Entry(root, width=30)
word_search.grid(row=3, column=2, columnspan=2)
word_search.insert(END, "단어를 입력하세요")

def add_file():     #해결
    files = filedialog.askopenfilenames(initialdir="/", \
        title="파일을 선택하세요",\
        filetypes=(("Excel 통합 문서", "*.xlsx"), ("모든 파일", "*.*"))) #무조건 튜플로 해야함
    for file in files:
        list_file.insert(END, file)
    print(files)
    # return files
def del_file():     #해결
    selected_file = list_file.curselection()
    list_file.delete(selected_file)
def trans_file():   #변환(엑셀파일->dictionnary)    #해결
    global groupe
    global length_word
    groupe.clear()
    length_word = 0
    selected_file= []
    index_selected_file = list_file.curselection()
    for i in index_selected_file:
        t = list_file.get(i)
        selected_file.append(t)
        #테스트
    for file in selected_file:
        file = file.replace("/", "\\")
        wb = load_workbook(file)
        ws = wb.active
        for word_row in ws.iter_rows(min_row=2, min_col=2, max_col=4, values_only=True):
            groupe.append(word_row)
        length_word = length_word+len(list(ws.rows))
    print(groupe)
    print(length_word)
    selected_file.clear()
    return groupe, length_word
def mean_list():    #해결
    for x in range(length_word-1):
        print(groupe[x][2])
def word_list():    #해결
    for x in range(length_word-1):
        print(groupe[x][:2])
def test_start_word():
    response = msgbox.askokcancel("확인 메세지", "불어 단어를 보고 뜻을 찾는 테스트를 하시겠습니까?")
    if response == 1 :
        new_window = Toplevel(root)
        new_window.title("불어 단어로 테스트")
        new_window.geometry()
        shuffle(groupe)

def test_start_mean():
    response = msgbox.askokcancel("확인 메세지", "뜻을 보고 단어를 찾는 테스트를 하시겠습니까?")
    if response == 1:
        shuffle(groupe)
        class Test:
            def __init__(self):
                self.new_window = Toplevel(root)
                self.new_window.title("뜻으로 테스트")
                self.new_window.geometry("480x300")
                self.btn_new_window_start = Button(self.new_window, text="테스트 시작", command=self.start_test)
                self.btn_new_window_start.pack()
                self.new_window.mainloop()
            def start_test(self):
                self.x = 0
                self.btn_new_window_start.destroy()
                self.display_word = StringVar()
                self.display_word.set("{0}".format(groupe[self.x][2]))
                self.word = Label(self.new_window, relief="groove", height=5, width=20, padx=10, pady=10, textvariable=self.display_word).pack()
                self.answer = Entry(self.new_window, width=30).pack()
                self.new_window.bind("<Return>", self.next_word)
            def next_word(self, event):
                self.x += 1
                self.display_word.set("{0}".format(groupe[self.x][2]))
            # def finish_test(self):

        app = Test()
            #     num = 0

            # # answer.grid(row=1, column=1, sticky=W+E)
            # def next_qt(event):
            #     num = 0
            #     word["text"] = "{0}".format(groupe[num][2])
            #     num += 1
            # new_window.bind("<Return>", next_qt)
            # word = Label(new_window, relief="groove", height=5, width=20, padx=10, pady=10, text="").pack()
            # answer = Entry(new_window, width=30).pack()
            #     # add_word = display_word.get()
            #     # add_word += "{0}".format(groupe[x][2])
            #     # display_word.set(add_word)
            #     # print(type(add_word))
            #     # word.grid(row=0, column=1, sticky=W+E)
            # if num == length_word-1:
            #     response1 = msgbox.askokcancel("종료", "테스트를 종료 및 결과를 보시겠습니까?")
            #     if response1 == 1:
            #         pass        

    #오늘은 여기까지. answer에서 값 입력 후 enter을 쳤을 때 다음으로 넘어가게 해주는 작업 해야 함.


            # for x in range(length_word-1):
            #     test_word = groupe[x][2]
            #     word = Label(new_window, text="{0}".format(test_word), font=("arial bold", 20), relief="groove", height=5, width=20, padx=10, pady=10)
            #     word.pack(new_window)
        
        

    

    # 백그라운드 모노톤으로 배경 디자인(간단하게만) (단어 테스트 앱 참고)
    # 엔터로 입력이 바로 될 수 있게 하기(다른 곳에도) 엔터 이후 바로 다음 문제로 넘어가기(애니메이션 관련 가능한지 확인, with 나도코딩 게임 개발 관련 영상 참고)
def cmd_mean_search():  #해결
    search = mean_search.get()
    num_row_mean = 0
    while num_row_mean<length_word-1:
        if search == groupe[num_row_mean][2]:
            print(groupe[num_row_mean][:2])
            break
        num_row_mean += 1
        if num_row_mean == length_word-1:
            print("찾는 단어가 없습니다")
            break
def cmd_word_search():  #해결
    search = word_search.get()
    num_row_word = 0
    while num_row_word<length_word-1:
        if search == groupe[num_row_word][1]:
            print(groupe[num_row_word][2])
            break
        num_row_word += 1
        if num_row_word == length_word-1:
            print("찾는 뜻이 없습니다")
            break

btn_add_file = Button(root, text="파일추가", command=add_file)
btn_del_file = Button(root, text="파일삭제", command=del_file)
btn_trans = Button(root, text="테스트할 단어파일 변환", command=trans_file)

# btn_mean_list = Button(root, text="파일 내 모든 뜻", command=mean_list)
btn_mean_search = Button(root, text="뜻으로 단어 찾기", command=cmd_mean_search)

# btn_word_list = Button(root, text="파일 내 모든 단어", command=word_list)
btn_word_search = Button(root, text="단어로 뜻 찾기", command=cmd_word_search)

btn_test_start_word = Button(root, text="불어 단어로 테스트 시작", command=test_start_word)
btn_test_start_mean = Button(root, text="한글 뜻으로 테스트 시작", command=test_start_mean)

btn_add_file.grid(row=0, column=2, sticky=W+E)
btn_del_file.grid(row=0, column=3, sticky=W+E)
btn_trans.grid(row=1, column=0, columnspan=4, sticky=W+E)

btn_mean_search.grid(row=4, column=0, columnspan=2, sticky=W+E)
# btn_mean_list.grid(row=2, column=0, columnspan=2, sticky=W+E)

btn_word_search.grid(row=4, column=2, columnspan=2, sticky=W+E)
# btn_word_list.grid(row=2, column=2, columnspan=2, sticky=W+E)

btn_test_start_word.grid(row=2, column=2, columnspan=2, sticky=W+E)
btn_test_start_mean.grid(row=2, column=0, columnspan=2, sticky=W+E)



root.mainloop()